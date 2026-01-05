import json
import asyncio
import time
import shutil
import urllib.parse
from contextlib import asynccontextmanager
from pathlib import Path
from datetime import datetime
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, UploadFile, File, Form, HTTPException, Request, Body
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# --- FIXED IMPORTS (Relative for 'app' folder) ---
from . import database
from .school_id_processor import SchoolIDProcessor, CONFIG

# ==================== WEBSOCKET MANAGER ====================
class ConnectionManager:
    def __init__(self): self.active_connections: list[WebSocket] = []
    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)
    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections: self.active_connections.remove(websocket)
    async def broadcast(self, message: dict):
        for connection in self.active_connections[:]:
            try: await connection.send_json(message)
            except: self.disconnect(connection)

manager = ConnectionManager()

# ==================== WATCHDOG ====================
class IDGenerationHandler(FileSystemEventHandler):
    def __init__(self, processor, loop):
        self.processor = processor
        self.loop = loop
        self.processing = set()

    def on_created(self, event):
        if event.is_directory: return
        filepath = event.src_path
        if not filepath.lower().endswith(('.jpg', '.jpeg', '.png')): return
        if filepath in self.processing: return
        self.processing.add(filepath)

        print(f"\nNew photo detected: {Path(filepath).name}")
        time.sleep(0.5) 
        
        try:
            success = self.processor.process_photo(filepath)
            if success:
                student_id = Path(filepath).stem
                front_file = f"{student_id}_FRONT.png"
                back_file = f"{student_id}_BACK.png"
                
                # Get student data from database
                student_data = database.get_student(student_id)
                
                msg = {
                    "type": "id_generated",
                    "data": {
                        "student_id": student_id,
                        "id_number": student_id,
                        "full_name": student_data.get('full_name', '') if student_data else '',
                        "section": student_data.get('section', '') if student_data else '',
                        "lrn": student_data.get('lrn', '') if student_data else '',
                        "front_url": f"/output/{urllib.parse.quote(front_file)}",
                        "back_url": f"/output/{urllib.parse.quote(back_file)}",
                        "front_image": f"/output/{urllib.parse.quote(front_file)}",
                        "back_image": f"/output/{urllib.parse.quote(back_file)}",
                        "timestamp": datetime.now().isoformat(),
                        "created_at": datetime.now().isoformat()
                    }
                }
                if self.loop and self.loop.is_running():
                    asyncio.run_coroutine_threadsafe(manager.broadcast(msg), self.loop)
                    print(f"✓ WebSocket broadcast sent for {student_id}")

        except Exception as e: print(f"CRITICAL ERROR: {e}")
        finally: self.processing.remove(filepath)

# ==================== APP LIFECYCLE ====================
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Ensure folders exist
    Path(CONFIG['OUTPUT_FOLDER']).mkdir(parents=True, exist_ok=True)
    Path(CONFIG['INPUT_FOLDER']).mkdir(parents=True, exist_ok=True)
    Path(CONFIG['TEMPLATE_FOLDER']).mkdir(parents=True, exist_ok=True)
    database.init_db()
    
    main_loop = asyncio.get_running_loop()
    processor = SchoolIDProcessor(CONFIG)
    event_handler = IDGenerationHandler(processor, main_loop)
    observer = Observer()
    observer.schedule(event_handler, CONFIG['INPUT_FOLDER'], recursive=False)
    observer.start()
    
    app.state.processor = processor
    app.state.observer = observer
    print(f"System Online: Watching {CONFIG['INPUT_FOLDER']}")
    yield
    app.state.observer.stop()
    app.state.observer.join()

app = FastAPI(lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# ==================== ROUTES ====================

@app.get("/api/layout")
def get_layout():
    try:
        path = CONFIG.get('LAYOUT_FILE', 'data/layout.json')
        with open(path, 'r') as f: return json.load(f)
    except: return {}

@app.post("/api/layout")
async def save_layout(request: Request):
    data = await request.json()
    path = CONFIG.get('LAYOUT_FILE', 'data/layout.json')
    with open(path, 'w') as f: json.dump(data, f, indent=4)
    return {"status": "saved"}

@app.get("/api/settings")
def get_settings():
    try:
        with open('data/settings.json', 'r') as f: return json.load(f)
    except: return {}

@app.post("/api/settings")
async def save_settings(request: Request):
    data = await request.json()
    with open('data/settings.json', 'w') as f: json.dump(data, f, indent=4)
    app.state.processor.reload_config() 
    return {"status": "saved"}

@app.get("/api/templates/list")
def list_templates():
    files = [f.name for f in Path(CONFIG['TEMPLATE_FOLDER']).glob("*.png")]
    return files

@app.get("/api/templates")
def get_templates():
    """Get list of templates with metadata - returns front/back structure"""
    all_templates = []
    for f in Path(CONFIG['TEMPLATE_FOLDER']).glob("*.png"):
        template_path = f"/templates/{urllib.parse.quote(f.name)}"
        all_templates.append({
            "id": f.stem,
            "name": f.stem,
            "filename": f.name,
            "path": template_path,  # Frontend uses 'path' property
            "url": template_path,
            "thumbnail": template_path,
            "size": f.stat().st_size,
            "modified": f.stat().st_mtime
        })
    
    # Separate into front and back based on naming convention
    # Templates with 'front' or '1' are front, others are back
    front_templates = [t for t in all_templates if 'front' in t['name'].lower() or t['name'] in ['1', 'rimberio_template', 'wardiere_template']]
    back_templates = [t for t in all_templates if 'back' in t['name'].lower() or t['name'] in ['2']]
    
    # If no specific naming, put all in both
    if not front_templates and not back_templates:
        front_templates = all_templates
        back_templates = all_templates
    
    return {
        "front": front_templates,
        "back": back_templates if back_templates else front_templates
    }

@app.post("/api/templates/upload")
async def upload_template(files: list[UploadFile] = File(...)):
    """Upload one or more template files"""
    uploaded = []
    try:
        for file in files:
            save_path = Path(CONFIG['TEMPLATE_FOLDER']) / file.filename
            with open(save_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            uploaded.append(file.filename)
        return {"status": "uploaded", "count": len(uploaded), "filenames": uploaded}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.delete("/api/templates/{filename}")
async def delete_template(filename: str):
    """Delete a template file"""
    template_path = Path(CONFIG['TEMPLATE_FOLDER']) / filename
    
    if not template_path.exists():
        raise HTTPException(status_code=404, detail="Template not found")
    
    try:
        template_path.unlink()
        return {"status": "deleted", "filename": filename}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/students")
def get_students(): return database.get_all_students()

@app.get("/api/students/search")
def search_students(q: str = ""):
    """Search students by name or ID"""
    query = q.strip()
    
    # Return empty for truly empty queries
    if not query or query == "":
        return []
    
    query_lower = query.lower()
    all_students = database.get_all_students()
    
    # Filter students by name or ID
    results = []
    for student in all_students:
        student_id = str(student.get('id_number') or student.get('student_id') or '').lower()
        student_name = str(student.get('full_name') or '').lower()
        
        if query_lower in student_id or query_lower in student_name:
            # Ensure student has id_number field for frontend compatibility
            if 'id_number' not in student and 'student_id' in student:
                student['id_number'] = student['student_id']
            results.append(student)
    
    return results[:10]  # Return top 10 matches

@app.put("/api/students/{student_id}")
async def update_student(student_id: str, data: dict = Body(...)):
    """Update student information"""
    conn = database.get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database connection failed")
    
    cursor = conn.cursor()
    sql = """
        UPDATE students 
        SET full_name=%s, lrn=%s, grade_level=%s, section=%s, 
            guardian_name=%s, address=%s, guardian_contact=%s 
        WHERE id_number=%s
    """
    val = (
        data.get('full_name', ''),
        data.get('lrn', ''),
        data.get('grade_level', ''),
        data.get('section', ''),
        data.get('guardian_name', ''),
        data.get('address', ''),
        data.get('guardian_contact', ''),
        student_id
    )
    cursor.execute(sql, val)
    conn.commit()
    conn.close()
    return {"status": "updated", "student_id": student_id}

@app.post("/api/students/update")
async def update_student(data: dict = Body(...)):
    conn = database.get_db_connection()
    cursor = conn.cursor()
    sql = """
        UPDATE students 
        SET full_name=%s, lrn=%s, grade_level=%s, section=%s, 
            guardian_name=%s, address=%s, guardian_contact=%s 
        WHERE id_number=%s
    """
    val = (
        data['name'], data['lrn'], data['grade_level'], data['section'], 
        data['guardian_name'], data['address'], data['guardian_contact'], 
        data['id']
    )
    cursor.execute(sql, val)
    conn.commit()
    conn.close()
    return {"status": "updated"}

@app.post("/api/regenerate/{student_id}")
async def regenerate_id(student_id: str):
    raw_path = Path(CONFIG['INPUT_FOLDER']) / f"{student_id}.jpg"
    if not raw_path.exists(): raw_path = Path(CONFIG['INPUT_FOLDER']) / f"{student_id}.png"
    
    if raw_path.exists():
        success = app.state.processor.process_photo(str(raw_path))
        return {"status": "regenerated" if success else "failed"}
    else:
        raise HTTPException(status_code=404, detail="Original photo not found")

@app.get("/api/history")
def get_history(limit: int = 50): 
    history = database.get_recent_history(limit=limit)
    enhanced_history = []
    for h in history:
        sid = h['student_id']
        # Frontend expects both field name variations
        front_path = f"/output/{sid}_FRONT.png"
        back_path = f"/output/{sid}_BACK.png"
        enhanced_history.append({
            **h,
            "id_number": h.get('student_id'),  # Add id_number alias
            "created_at": h.get('timestamp'),  # Add created_at alias
            "front_url": front_path,
            "back_url": back_path,
            "front_image": front_path,  # Frontend uses this
            "back_image": back_path    # Frontend uses this
        })
    return enhanced_history

@app.get("/api/system/stats")
def get_system_stats():
    """Get system statistics"""
    students = database.get_all_students()
    history = database.get_recent_history(limit=1000)
    
    output_path = Path(CONFIG['OUTPUT_FOLDER'])
    input_path = Path(CONFIG['INPUT_FOLDER'])
    
    # Count files
    output_files = list(output_path.glob("*.png"))
    input_files = list(input_path.glob("*.[jp][pn]g"))
    
    return {
        "total_students": len(students),
        "total_generated": len(history),
        "output_files": len(output_files),
        "input_files": len(input_files),
        "templates": len(list(Path(CONFIG['TEMPLATE_FOLDER']).glob("*.png"))),
        "disk_usage": {
            "output_mb": sum(f.stat().st_size for f in output_files) / (1024 * 1024),
            "input_mb": sum(f.stat().st_size for f in input_files) / (1024 * 1024)
        }
    }

# --- UPDATED CAPTURE ENDPOINT ---
@app.post("/api/capture")
async def upload_capture(
    file: UploadFile = File(...), 
    student_id: str = Form(...),
    # Manual Fields
    manual_name: str = Form(None),
    manual_grade: str = Form(None),
    manual_section: str = Form(None),
    manual_guardian: str = Form(None), # NEW
    manual_address: str = Form(None),  # NEW
    manual_contact: str = Form(None)   # NEW
):
    try:
        # 1. If manual data exists, save it to a JSON sidecar file
        if manual_name:
            json_path = Path(CONFIG['INPUT_FOLDER']) / f"{student_id}.json"
            manual_data = {
                "id_number": student_id,
                "full_name": manual_name,
                "grade_level": manual_grade or "",
                "section": manual_section or "",
                # Capture the new Back of ID fields
                "guardian_name": manual_guardian or "", 
                "address": manual_address or "", 
                "guardian_contact": manual_contact or "",
                "lrn": "" # Default empty
            }
            with open(json_path, 'w') as f:
                json.dump(manual_data, f)
            print(f"Manual data saved for {student_id}")

        # 2. Save the image (Triggers Watchdog)
        filename = f"{student_id}.jpg"
        save_path = Path(CONFIG['INPUT_FOLDER']) / filename
        with open(save_path, "wb") as buffer: 
            shutil.copyfileobj(file.file, buffer)
            
        return {"status": "saved", "path": str(save_path)}
    except Exception as e: 
        print(f"Capture Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/analytics/export")
def export_analytics():
    """Export analytics data as JSON"""
    students = database.get_all_students()
    history = database.get_recent_history(limit=1000)
    
    # Group by date
    from collections import defaultdict
    daily_stats = defaultdict(int)
    
    for h in history:
        date = h['timestamp'].strftime('%Y-%m-%d') if h.get('timestamp') else 'unknown'
        daily_stats[date] += 1
    
    return {
        "total_students": len(students),
        "total_generated": len(history),
        "daily_generation": dict(sorted(daily_stats.items())),
        "recent_history": history[:50],
        "export_date": datetime.now().isoformat()
    }

@app.post("/api/students/import")
async def import_students(file: UploadFile = File(...)):
    """Import students from CSV or Excel file"""
    import csv
    import io
    
    if not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")
    
    try:
        content = await file.read()
        
        # Parse CSV
        if file.filename.endswith('.csv'):
            text = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
        else:
            # Parse Excel
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            except ImportError:
                raise HTTPException(status_code=400, detail="Excel support requires openpyxl package")
        
        # Validate required columns
        required_cols = ['ID_Number', 'Full_Name', 'LRN', 'Section', 'Guardian_Name', 'Address', 'Guardian_Contact']
        if not all(col in rows[0].keys() for col in required_cols):
            raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(required_cols)}")
        
        # Import to database
        conn = database.get_db_connection()
        if not conn:
            raise HTTPException(status_code=500, detail="Database connection failed")
        
        cursor = conn.cursor()
        imported = 0
        errors = []
        
        for idx, row in enumerate(rows, start=2):
            try:
                sql = """
                    INSERT INTO students 
                    (id_number, full_name, lrn, grade_level, section, guardian_name, address, guardian_contact) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s) 
                    ON DUPLICATE KEY UPDATE 
                    full_name = VALUES(full_name), 
                    lrn = VALUES(lrn),
                    grade_level = VALUES(grade_level), 
                    section = VALUES(section),
                    guardian_name = VALUES(guardian_name),
                    address = VALUES(address),
                    guardian_contact = VALUES(guardian_contact)
                """
                val = (
                    str(row.get('ID_Number', '')).strip(),
                    str(row.get('Full_Name', '')).strip(),
                    str(row.get('LRN', '')).strip(),
                    str(row.get('Grade_Level', '')).strip(),
                    str(row.get('Section', '')).strip(),
                    str(row.get('Guardian_Name', '')).strip(),
                    str(row.get('Address', '')).strip(),
                    str(row.get('Guardian_Contact', '')).strip()
                )
                cursor.execute(sql, val)
                imported += 1
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})
        
        conn.commit()
        conn.close()
        
        return {
            "status": "success",
            "imported": imported,
            "total": len(rows),
            "errors": errors[:10]  # Return first 10 errors
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

@app.post("/api/students/import/preview")
async def preview_import(file: UploadFile = File(...)):
    """Preview CSV/Excel file before importing"""
    import csv
    import io
    
    if not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")
    
    try:
        content = await file.read()
        
        # Parse file
        if file.filename.endswith('.csv'):
            text = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
        else:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):  # Skip empty rows
                        rows.append(dict(zip(headers, row)))
            except ImportError:
                raise HTTPException(status_code=400, detail="Excel support requires openpyxl package")
        
        # Validate columns
        required_cols = ['ID_Number', 'Full_Name', 'LRN', 'Section', 'Guardian_Name', 'Address', 'Guardian_Contact']
        headers = list(rows[0].keys()) if rows else []
        missing_cols = [col for col in required_cols if col not in headers]
        
        return {
            "total_rows": len(rows),
            "headers": headers,
            "required_columns": required_cols,
            "missing_columns": missing_cols,
            "preview_data": rows[:10],  # First 10 rows
            "valid": len(missing_cols) == 0
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Preview failed: {str(e)}")

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True: await websocket.receive_text()
    except WebSocketDisconnect: manager.disconnect(websocket)

app.mount("/output", StaticFiles(directory=CONFIG['OUTPUT_FOLDER']), name="output")
app.mount("/templates", StaticFiles(directory=CONFIG['TEMPLATE_FOLDER']), name="templates")